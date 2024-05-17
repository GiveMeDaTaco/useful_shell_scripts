import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Sample condition_dict and DataFrames for demonstration
condition_dict = {
    "Condition 1": ["Logic 1", "Description 1"],
    "Condition 2": ["Logic 2", "Description 2"],
    "Condition 3": ["Logic 3", "Description 3"]
}

# Sample DataFrames for dynamic population with user-defined identifiers
identifier_1 = "ID1"
identifier_2 = "ID2"

df1 = pd.DataFrame({
    f'{identifier_1} drop if only this scrub': [1, 2, 3],
    f'{identifier_1} drop incremental': [4, 5, 6],
    f'{identifier_1} drop cumulative': [7, 8, 9],
    f'{identifier_1} party regain if not scrub': [10, 11, 12],
    f'{identifier_1} remaining': [13, 14, 15]
})

df2 = pd.DataFrame({
    f'{identifier_2} drop if only this scrub': [16, 17, 18],
    f'{identifier_2} drop incremental': [19, 20, 21],
    f'{identifier_2} drop cumulative': [22, 23, 24],
    f'{identifier_2} party regain if not scrub': [25, 26, 27],
    f'{identifier_2} remaining': [28, 29, 30]
})

dataframes = [df1, df2]

# User-provided details
cp_name = "John Doe"
lead_name = "Jane Smith"
current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Set offer_code in A1 with the required format, bold and font size 14
offer_code = "OFFER2024"
ws['A1'] = f'{offer_code} [{current_date}] [CP: {cp_name}] [Lead: {lead_name}]'
ws['A1'].font = Font(bold=True, size=14)

# Set columns A2:C2 with background color, text alignment, wrapping, and bold font
header = ["Check #", "Criteria Logic", "Criteria Description"]
for col_num, header_title in enumerate(header, 1):
    cell = ws.cell(row=2, column=col_num, value=header_title)
    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.font = Font(bold=True)

# Set row 2 height
ws.row_dimensions[2].height = 40

# Set A3 to "Starting Population"
ws['A3'] = "Starting Population"

# Populate rows A4 - Cn with condition_dict
for row_num, (key, values) in enumerate(condition_dict.items(), start=4):
    ws[f'A{row_num}'] = key
    ws[f'B{row_num}'] = values[0]
    ws[f'C{row_num}'] = values[1]

# Set initial column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 53
ws.column_dimensions['C'].width = 65

# Dynamically populate columns E4 onward with pandas dataframes
col_start = 5  # Start from column E
for df in dataframes:
    # Set divider column width and dark gray background
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_start - 1)].width = 3
    for row in range(2, len(condition_dict) + 4):
        cell = ws.cell(row=row, column=col_start - 1)
        cell.fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")

    # Populate dataframe columns
    for col_num, column_title in enumerate(df.columns, col_start):
        ws.cell(row=2, column=col_num, value=column_title)
        cell = ws.cell(row=2, column=col_num)
        cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.font = Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15
        for row_num, value in enumerate(df[column_title], start=4):
            ws.cell(row=row_num, column=col_num, value=value)
    col_start += len(df.columns) + 1  # Move to the next group of columns

# Save the workbook
wb.save("/mnt/data/offer_conditions_final.xlsx")
